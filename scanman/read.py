from lxml import etree
from .model import Vulnerability, Host
from openpyxl import load_workbook
import pandas as pd
import json
import re


class Parser:
  def __init__(self) -> None:
    pass

  def clean(self, text: str):
    ret = text
    bad_tokens = ['\n', '\t', ' ']
    for token in bad_tokens:
      ret = ret.replace(token, "")
    return ret

  def detect(self, text):
    """
    检测文本是否为该Parser可处理的格式
    子类可以重写此方法来实现特定的指纹识别
    
    Args:
        text: 输入的HTML文本内容（bytes或str）
    
    Returns:
        bool: True表示该Parser可以处理此格式，False表示不能处理
    """
    return False

  def parse_vulnerability(self, text):
    pass

  def parse_host(self, text):
    pass


class RSASParser(Parser):
  def detect(self, text):
    """检测是否为RSAS格式的报告"""
    if isinstance(text, bytes):
      try:
        text = text.decode('utf-8')
      except UnicodeDecodeError:
        text = text.decode('gbk', errors='ignore')
    
    # RSAS特征：包含特定的HTML结构
    rsas_patterns = [
      '<h1>绿盟科技&#34;远程安全评估系统&#34;安全评估报告-主机报表</h1>',
      'media/report/js/nsfocus',
    ]
    
    return all(pattern in text for pattern in rsas_patterns)

  def parse_vulnerability(self, text):
    root = etree.HTML(text)

    vuln_names = root.xpath(
        '//*[@id="vul_detail"]/table/tr/td/span/text()')
    vuln_names = list(map(str.strip, vuln_names))

    vuln_threat = root.xpath(
        '//*[@id="vul_detail"]/table/tr/td/span/@class')
    vuln_threat = list(map(lambda x: x.split('_')[2], vuln_threat))

    nodes = root.xpath('//*[@class="solution"]/td/table/tr[1]/td')
    vuln_desc = ["".join(node.itertext()) for node in nodes]
    vuln_desc = list(map(self.clean, vuln_desc))

    nodes = root.xpath('//*[@class="solution"]/td/table/tr[2]/td')
    vuln_solution = ["".join(node.itertext()) for node in nodes]
    vuln_solution = list(map(self.clean, vuln_solution))

    fields = [vuln_names, vuln_threat, vuln_solution, vuln_desc]
    assert (any(len(field) == len(fields[0]) for field in fields))

    ret = []

    for i in range(len(vuln_names)):
      ret.append(Vulnerability(
          name=vuln_names[i],
          severity=vuln_threat[i],
          description=vuln_desc[i],
          solution=vuln_solution[i]
      ))
    return ret

  def parse_host(self, text):
    root = etree.HTML(text)

    host_ip = root.xpath(
        '//*[@id="content"]/div[2]/table[2]/tr/td[1]/table/tbody/tr[1]/td/text()')[0]
    vuln_names = root.xpath(
        '//*[@id="vul_detail"]/table/tr/td/span/text()')
    vuln_names = list(map(str.strip, vuln_names))

    ports = root.xpath('/html/body/div/div[4]/div[6]/div[2]/table/tbody/tr/td[1]/text()')
    ports = [_.strip() for _ in ports]

    return Host(ip=host_ip, name="", ports=ports), vuln_names


class TRXParser(Parser):
  def detect(self, text):
    """检测是否为TRX格式的报告"""
    if isinstance(text, bytes):
      try:
        text = text.decode('utf-8')
      except UnicodeDecodeError:
        text = text.decode('gbk', errors='ignore')
    
    # TRX特征：这里需要根据实际的TRX报告格式来定义
    # 暂时返回False，等待具体实现
    return False

  def parse_vulnerability(self, text):
    return super().parse_vulnerability(text)

  def parse_host(self, text):
    return super().parse_host(text)


class XLSXParser:
  def __init__(self) -> None:
    pass

  def read_host_name_ip(self, path):
    wb = load_workbook(path)
    # ws = list(wb)[0]

    ret = {}
    for ws in list(wb):
      # ip_col = -1
      name_col = -1
      row = 1
      ip_cols = []
      for col in range(1, 64):
        value = ws.cell(column=col, row=row).value
        if value == 'ip':
          ip_col = col
          ip_cols.append(col)
        if value == 'name':
          name_col = col
      row = 2
      while True:
        if ws.cell(column=ip_col, row=row).value == None:
          break
        # ip = ws.cell(column=ip_col, row=row).value
        ips = [ws.cell(column=col, row=row).value for col in ip_cols]
        name = ws.cell(column=name_col, row=row).value
        for ip in ips:
          ret[ip] = name
        # ret[ip] = ''
        # ret[ip] += name
        row += 1
    return ret

class XLSXReportParser(Parser):
  def parse(self, path):
    wb = load_workbook(path)
    ws = list(wb)[0]
    host_ip_col = 2
    host_name_col = 3
    vulnerability_name_col = 14
    vulnerability_severity_col = 13
    vulnerability_decription_col = 20
    vulnerability_solution_col = 21

    row = 2

    def get_col_value(col):
      return ws.cell(column=col, row=row).value
    
    hosts = []
    vuls = []
    affections = {}

    while True:
      host_name = get_col_value(col=host_name_col)
      host_ip = get_col_value(col=host_ip_col)
      vul_name = get_col_value(col=vulnerability_name_col)
      vul_severity = get_col_value(col=vulnerability_severity_col)
      vul_description = get_col_value(col=vulnerability_decription_col)
      vul_solution = get_col_value(col=vulnerability_solution_col)
      row += 1
      severity_map = {'高危':'high','中危':'middle','低危':'low','危急':'critical'}
      if not host_ip:
        break
      vul = Vulnerability(name=vul_name, severity=severity_map[vul_severity], description=vul_description, solution=vul_solution)
      host = Host(name=host_name, ip=host_ip)

      if vul not in vuls:
        vuls.append(vul)
      if host not in hosts:
        hosts.append(host)
      if vul.name not in affections:
        affections[vul.name] = []
      affections[vul.name].append(host.ip)
    return vuls, hosts, affections


    pass

  def parse_host(self, path):
    pass

class XLSXSelectiveRemoveParser(Parser):
  def parse(self, path):
    df = pd.read_excel(path)

    # 确定文件中是否有我们需要的列名，例如vuln_name和ip，这里假设这两列存在
    # 并将这些数据转换为二元组列表
    tuple_list = list(zip(df['vuln_name'], df['ip']))

    # 显示前几个元素以验证
    return tuple_list
  
  def function_parse(path):
    selective_parser = XLSXSelectiveRemoveParser()
    return selective_parser.parse(path=path)
   


class WANGSHENParser(Parser):
  def detect(self, text):
    """检测是否为网神格式的报告"""
    if isinstance(text, bytes):
      try:
        text = text.decode('utf-8')
      except UnicodeDecodeError:
        text = text.decode('gbk', errors='ignore')
    
    # 网神特征：包含特定的CSS类名
    wangshen_patterns = [
      'class="odd vuln_middle"',
      'class="more hide odd"',
      'class="report_table plumb"'
    ]
    
    return all(pattern in text for pattern in wangshen_patterns)

  def parse_vulnerability(self, text):
    root = etree.HTML(text)

    name_nodes = root.xpath('//*[@class="odd vuln_middle"]')
    nodes = root.xpath('//*[@class="more hide odd"]')
    
    ret = []

    for i in range(len(nodes)):
      node = nodes[i]
      name = name_nodes[i].xpath('td[1]/span/text()')[0]
      threat = node.xpath('td/table/tr[2]/td/text()')[0][0]
      threat_map = {
        "高":"high",
        "中":"middle",
        "低":"low"
      }
      threat = threat_map[threat]
      desc = node.xpath('td/table/tr[4]/td/text()')[0]
      try:
        solution = node.xpath('td/table/tr[5]/td/text()')[0]
      except:
        solution = ""
      new_vuln = Vulnerability(
        name=name,
        severity = threat,
        description = desc,
        solution=solution
      )
      ret.append(new_vuln)
    return ret
  
  def parse_host(self, text):
    root = etree.HTML(text)

    host_ip = root.xpath('//*[@class="report_table plumb"]/tbody/tr[2]/td/text()')[0]
    vuln_names = root.xpath('//*[@class="odd vuln_middle"]/td[1]/span/text()')
    return Host(ip=host_ip), vuln_names

class NSFOCUSParser(Parser):
  def detect(self, text):
    """检测是否为绿盟NSFOCUS格式的报告"""
    if isinstance(text, bytes):
      try:
        text = text.decode('utf-8')
      except UnicodeDecodeError:
        text = text.decode('gbk', errors='ignore')
    
    # NSFOCUS特征：包含window.data变量定义
    nsfocus_patterns = [
      'window.data',
      '"mark":"vul-detail"',
      '"mark":"host-summary"'
    ]
    
    return all(pattern in text for pattern in nsfocus_patterns)

  def _extract_json(self, text):
    # The data is in a script tag as a javascript variable.
    # We can use regex to extract it.
    if isinstance(text, bytes):
      try:
        text = text.decode('utf-8')
      except UnicodeDecodeError:
        text = text.decode('gbk', errors='ignore')
        
    match = re.search(r'window\.data\s*=\s*(\{.*?\});', text, re.DOTALL)
    if match:
      json_str = match.group(1)
      return json.loads(json_str)
    return None

  def parse_vulnerability(self, text):
    data = self._extract_json(text)
    if not data:
      return []

    ret = []
    vuln_info_category = next((cat for cat in data.get('categories', []) if cat.get('mark') == 'vul-detail'), None)
    
    if not vuln_info_category:
        return []

    for item in vuln_info_category.get('data', {}).get('vul_items', []):
      for vul in item.get('vuls', []):
        vul_msg = vul.get('vul_msg', {})
        name = vul_msg.get('i18n_name')
        severity = vul.get('vul_level') # 'low', 'middle', 'high'
        description = "".join(vul_msg.get('i18n_description', []))
        solution = "".join(vul_msg.get('i18n_solution', []))

        if name and severity:
          new_vuln = Vulnerability(
            name=name,
            severity=severity,
            description=self.clean(description),
            solution=self.clean(solution)
          )
          if new_vuln not in ret:
            ret.append(new_vuln)
            
    return ret

  def parse_host(self, text):
    data = self._extract_json(text)
    if not data:
      return None, []

    host_ip = ""
    host_name = ""
    ports = []
    vuln_names = []

    # Get host IP and name
    host_summary_category = next((cat for cat in data.get('categories', []) if cat.get('mark') == 'host-summary'), None)
    if host_summary_category:
      host_data = host_summary_category.get('data', {})
      host_ip = host_data.get('target', '')
      host_name = host_data.get('hostname', '')

    # Get ports
    other_info_category = next((cat for cat in data.get('categories', []) if cat.get('mark') == 'other-info'), None)
    if other_info_category:
        other_info_data = other_info_category.get('data',{}).get('other_info_data',[])
        remote_ports_info = next((info for info in other_info_data if info.get('info_name') == '远程端口信息'), None)
        if remote_ports_info:
            ports = [str(content[0]) for content in remote_ports_info.get('content', [])]


    # Get vulnerability names
    vuln_info_category = next((cat for cat in data.get('categories', []) if cat.get('mark') == 'vul-detail'), None)
    if vuln_info_category:
      for item in vuln_info_category.get('data', {}).get('vul_items', []):
        for vul in item.get('vuls', []):
          vul_msg = vul.get('vul_msg', {})
          name = vul_msg.get('i18n_name')
          if name and name not in vuln_names:
            vuln_names.append(name)
    
    host = Host(ip=host_ip, name=host_name, ports=ports)
    return host, vuln_names

class GreenLeagueParser(Parser):
  def detect(self, text):
    """检测是否为绿盟GreenLeague格式的报告"""
    if isinstance(text, bytes):
      try:
        text = text.decode('utf-8')
      except UnicodeDecodeError:
        text = text.decode('gbk', errors='ignore')
    
    # GreenLeague特征：包含特定的script标签结构
    greenleague_patterns = [
      '<script>window.data = ',
    ]
    
    return all(pattern in text for pattern in greenleague_patterns)

  def _extract_json(self, text):
    # The data is in a script tag as a javascript variable.
    # Use script tags as delimiters to extract JSON data.
    if isinstance(text, bytes):
      try:
        text = text.decode('utf-8')
      except UnicodeDecodeError:
        text = text.decode('gbk', errors='ignore')
    
    # Find the start and end positions
    start_marker = '<script>window.data = '
    end_marker = '</script>'
    
    start_pos = text.find(start_marker)
    if start_pos == -1:
      return None
    
    start_pos += len(start_marker)
    end_pos = text.find(end_marker, start_pos)
    if end_pos == -1:
      return None
    
    # Extract the JSON string and remove the trailing semicolon
    json_str = text[start_pos:end_pos].rstrip(';')
    try:
      return json.loads(json_str)
    except json.JSONDecodeError:
      return None

  def parse_vulnerability(self, text):
    data = self._extract_json(text)
    if not data:
      return []

    ret = []
    
    # 查找漏洞详情类别，可能在顶级类别或子类别中
    vuln_detail_data = None
    
    for category in data.get('categories', []):
      # 检查顶级类别
      if category.get('mark') == 'vul-detail':
        vuln_detail_data = category.get('data', {})
        break
      
      # 检查子类别
      if 'children' in category:
        for child in category['children']:
          if child.get('mark') == 'vul-detail':
            vuln_detail_data = child.get('data', {})
            break
        if vuln_detail_data:
          break
    
    if not vuln_detail_data:
      return []

    for item in vuln_detail_data.get('vul_items', []):
      for vul in item.get('vuls', []):
        vul_msg = vul.get('vul_msg', {})
        name = vul_msg.get('i18n_name')
        severity = vul.get('vul_level') # 'low', 'middle', 'high'
        description = "".join(vul_msg.get('i18n_description', []))
        solution = "".join(vul_msg.get('i18n_solution', []))

        if name and severity:
          new_vuln = Vulnerability(
            name=name,
            severity=severity,
            description=self.clean(description),
            solution=self.clean(solution)
          )
          if new_vuln not in ret:
            ret.append(new_vuln)
            
    return ret

  def parse_host(self, text):
    data = self._extract_json(text)
    if not data:
      return None, []

    host_ip = ""
    host_name = ""
    ports = []
    vuln_names = []

    # Get host IP and name
    host_summary_category = next((cat for cat in data.get('categories', []) if cat.get('mark') == 'host-summary'), None)
    if host_summary_category:
      host_data = host_summary_category.get('data', {})
      host_ip = host_data.get('target', '')
      host_name = host_data.get('hostname', '')

    # Get ports from other-info section
    other_info_category = next((cat for cat in data.get('categories', []) if cat.get('mark') == 'other-info'), None)
    if other_info_category:
        other_info_data = other_info_category.get('data', {}).get('other_info_data', [])
        remote_ports_info = next((info for info in other_info_data if info.get('info_name') == '远程端口信息'), None)
        if remote_ports_info:
            ports = [str(content[0]) for content in remote_ports_info.get('content', [])]

    # Get vulnerability names - 查找漏洞详情类别，可能在顶级类别或子类别中
    vuln_detail_data = None
    for category in data.get('categories', []):
      # 检查顶级类别
      if category.get('mark') == 'vul-detail':
        vuln_detail_data = category.get('data', {})
        break
      
      # 检查子类别
      if 'children' in category:
        for child in category['children']:
          if child.get('mark') == 'vul-detail':
            vuln_detail_data = child.get('data', {})
            break
        if vuln_detail_data:
          break
    
    if vuln_detail_data:
      for item in vuln_detail_data.get('vul_items', []):
        for vul in item.get('vuls', []):
          vul_msg = vul.get('vul_msg', {})
          name = vul_msg.get('i18n_name')
          if name and name not in vuln_names:
            vuln_names.append(name)
    
    host = Host(ip=host_ip, name=host_name, ports=ports)
    return host, vuln_names
